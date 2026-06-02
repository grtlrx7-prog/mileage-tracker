from openpyxl import load_workbook

FILE = "templates/SARS_Logbook_2026.xlsx"

workbook = load_workbook(FILE)

print(workbook.sheetnames)